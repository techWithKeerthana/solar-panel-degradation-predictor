"""
Report generator for prediction results.

Produces two export formats from a PredictionResult record:
  - PDF via fpdf2: one-page report with all inputs, prediction, and recommendation
  - Excel via openpyxl: two-sheet workbook (summary + full input table)

Both functions accept a PredictionResult object (ORM model) and return
a bytes object so the route can stream it directly without writing to disk.

Why fpdf2 instead of reportlab?
  fpdf2 is a simpler API for text-layout PDFs, has no external binary
  dependencies, and is sufficient for a structured one-page report.
  reportlab is more powerful but adds complexity that isn't needed here.
"""

from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _safe(text: str) -> str:
    """
    Replace Unicode characters that fall outside latin-1 (the encoding used
    by fpdf2's built-in core fonts) with safe ASCII equivalents.
    This avoids FPDFUnicodeEncodingException without requiring a bundled TTF.
    """
    replacements = {
        '\u2014': ' - ',   # em dash  ->  hyphen-space
        '\u2013': '-',     # en dash
        '\u2018': "'",     # left single quote
        '\u2019': "'",     # right single quote
        '\u201c': '"',     # left double quote
        '\u201d': '"',     # right double quote
        '\u2026': '...',   # ellipsis
        '\u00b0': 'deg',   # degree sign
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    return text


SEVERITY_COLORS = {
    # (R, G, B) tuples used in both PDF and Excel
    'critical': (198, 40, 40),    # deep red
    'warning':  (230, 81, 0),     # deep orange
    'ok':       (46, 125, 50),    # dark green
}

THEME_COLORS_RGB = {
    'navy': (15, 23, 41),          # #0F1729
    'offwhite': (241, 245, 249),   # #F1F5F9
    'amber': (245, 158, 11),       # #F59E0B
    'blue': (59, 130, 246),        # #3B82F6
    'green': (34, 197, 94),        # #22C55E
    'red': (239, 68, 68),          # #EF4444
    'muted': (148, 163, 184),      # #94A3B8
    'tint_row': (241, 245, 251),   # light navy tint for row striping
}

THEME_COLORS_HEX = {
    'navy': '0F1729',
    'offwhite': 'F1F5F9',
    'amber': 'F59E0B',
    'blue': '3B82F6',
    'green': '22C55E',
    'red': 'EF4444',
    'muted': '94A3B8',
    'tint_row_a': 'F8FAFC',
    'tint_row_b': 'EEF2F9',
}

SEVERITY_TINT_RGB = {
    'critical': (255, 239, 239),
    'warning':  (255, 247, 230),
    'ok':       (238, 252, 242),
}

SEVERITY_TINT_HEX = {
    'critical': 'FDECEC',
    'warning':  'FFF6DE',
    'ok':       'ECFDF3',
}

SEVERITY_HEX = {
    'critical': 'C62828',
    'warning':  'E65100',
    'ok':       '2E7D32',
}

def _get_severity(efficiency: float) -> str:
    """Classify efficiency into a severity level matching maintenance_rules.py."""
    if efficiency < 0.40:
        return 'critical'
    if efficiency < 0.55:
        return 'warning'
    return 'ok'


# ── PDF report ─────────────────────────────────────────────────────────────────

class _SolarPDF(FPDF):
    """Custom FPDF subclass with header and footer."""

    def header(self):
        # Title bar
        self.set_fill_color(*THEME_COLORS_RGB['navy'])
        self.rect(0, 0, 210, 22, 'F')
        self.set_font('Helvetica', 'B', 14)
        self.set_text_color(*THEME_COLORS_RGB['offwhite'])
        self.set_xy(10, 5)
        self.cell(0, 12, 'Solar Panel Degradation Predictor', align='L')
        self.ln(20)
        self.set_text_color(0, 0, 0)

    def footer(self):
        self.set_y(-12)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(*THEME_COLORS_RGB['muted'])
        self.cell(0, 10,
                  f'VTU BE CSE (AI&ML) - Rajeev Institute of Technology, Hassan  |  '
                  f'Page {self.page_no()}  |  Generated {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}',
                  align='C')


def generate_pdf(result) -> bytes:
    """
    Generate a one-page PDF report for a PredictionResult.

    Layout:
      - Header (branded bar)
      - Report metadata (result ID, date, user)
      - Predicted efficiency (large, colour-coded)
      - Maintenance recommendation box (colour-coded border)
      - Input features table (two-column layout)
      - Model note footer
      - Page footer

    Args:
        result: PredictionResult ORM object

    Returns:
        bytes: PDF file content, ready to stream as a download
    """
    pdf = _SolarPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    input_data = result.get_input_data()
    efficiency = result.predicted_efficiency
    severity = _get_severity(efficiency)
    sev_r, sev_g, sev_b = SEVERITY_COLORS[severity]

    # ── Section 1: Report metadata ─────────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(*THEME_COLORS_RGB['amber'])
    pdf.cell(0, 8, 'Prediction Report', ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(*THEME_COLORS_RGB['muted'])
    pdf.cell(0, 5, f'Result ID: #{result.id}', ln=True)
    pdf.cell(0, 5, f'Generated: {result.created_at.strftime("%Y-%m-%d %H:%M:%S UTC")}', ln=True)
    pdf.set_text_color(*THEME_COLORS_RGB['offwhite'])
    pdf.ln(4)

    # ── Section 2: Predicted efficiency (large coloured box) ──────────────────
    pdf.set_fill_color(*SEVERITY_TINT_RGB[severity])
    pdf.rect(10, pdf.get_y(), 190, 28, 'F')
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(*THEME_COLORS_RGB['muted'])
    pdf.set_x(10)
    pdf.cell(0, 10, 'Predicted Solar Panel Efficiency', ln=True)
    pdf.set_font('Helvetica', 'B', 28)
    pdf.set_text_color(sev_r, sev_g, sev_b)
    pdf.set_x(10)
    pdf.cell(0, 14, f'{efficiency:.4f}  ({efficiency * 100:.1f}%)', ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── Section 3: Maintenance recommendation ─────────────────────────────────
    # Coloured left border drawn as a filled rectangle
    bar_y = pdf.get_y()
    pdf.set_fill_color(sev_r, sev_g, sev_b)
    pdf.rect(10, bar_y, 4, 16, 'F')

    pdf.set_fill_color(*SEVERITY_TINT_RGB[severity])
    pdf.rect(14, bar_y, 186, 16, 'F')
    pdf.set_xy(16, bar_y + 2)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(sev_r, sev_g, sev_b)
    pdf.cell(0, 6, 'Maintenance Recommendation', ln=True)
    pdf.set_x(16)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(*THEME_COLORS_RGB['navy'])
    pdf.cell(0, 6, _safe(result.maintenance_recommendation), ln=True)
    pdf.set_text_color(*THEME_COLORS_RGB['offwhite'])
    pdf.ln(8)

    # ── Section 4: Input features table ───────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*THEME_COLORS_RGB['amber'])
    pdf.cell(0, 8, 'Input Feature Values', ln=True)

    pdf.set_font('Helvetica', '', 8)
    col_w = 90
    row_h = 6
    items = list(input_data.items())
    mid = (len(items) + 1) // 2    # split into two columns
    left_items = items[:mid]
    right_items = items[mid:]

    # Table header row
    pdf.set_fill_color(*THEME_COLORS_RGB['navy'])
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(*THEME_COLORS_RGB['offwhite'])
    pdf.cell(col_w, row_h, '  Feature', border=1, fill=True)
    pdf.cell(col_w, row_h, '  Feature', border=1, fill=True, ln=True)

    # Alternate row shading
    for i, (left, right) in enumerate(zip(left_items, right_items)):
        fill = THEME_COLORS_RGB['tint_row'] if i % 2 == 0 else THEME_COLORS_RGB['offwhite']
        pdf.set_fill_color(*fill)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*THEME_COLORS_RGB['navy'])
        lname, lval = left
        rname, rval = right
        lval_str = f'{lval:.4f}' if isinstance(lval, float) else str(lval)
        rval_str = f'{rval:.4f}' if isinstance(rval, float) else str(rval)
        pdf.cell(col_w // 2, row_h, f'  {lname}', border='LB', fill=True)
        pdf.cell(col_w // 2, row_h, lval_str, border='RB', fill=True)
        pdf.cell(col_w // 2, row_h, f'  {rname}', border='LB', fill=True)
        pdf.cell(col_w // 2, row_h, rval_str, border='RB', fill=True, ln=True)

    # Handle odd row if left column has one more entry
    if len(left_items) > len(right_items):
        lname, lval = left_items[-1]
        lval_str = f'{lval:.4f}' if isinstance(lval, float) else str(lval)
        fill = THEME_COLORS_RGB['tint_row'] if len(right_items) % 2 == 0 else THEME_COLORS_RGB['offwhite']
        pdf.set_fill_color(*fill)
        pdf.set_text_color(*THEME_COLORS_RGB['navy'])
        pdf.cell(col_w // 2, row_h, f'  {lname}', border='LB', fill=True)
        pdf.cell(col_w // 2, row_h, lval_str, border='RB', fill=True)
        pdf.cell(col_w, row_h, '', border='LRB', fill=True, ln=True)

    pdf.ln(6)

    # ── Section 5: Model note ──────────────────────────────────────────────────
    pdf.set_font('Helvetica', 'I', 8)
    pdf.set_text_color(*THEME_COLORS_RGB['muted'])
    pdf.multi_cell(0, 5,
        _safe('Model: Gradient Boosting Regressor (R2=0.790, RMSE=0.048, test set 4,000 samples). '
        'Primary driver: solar irradiance (feature importance 0.674). '
        'Maintenance thresholds derived from EDA (dataset mean efficiency ~0.52).'))

    return bytes(pdf.output())


# ── Excel report ───────────────────────────────────────────────────────────────

def generate_excel(result) -> bytes:
    """
    Generate a two-sheet Excel workbook for a PredictionResult.

    Sheet 1 - Summary: key result info, efficiency, recommendation, model note.
    Sheet 2 - Input Data: one column per feature with header + value rows.

    Args:
        result: PredictionResult ORM object

    Returns:
        bytes: .xlsx file content, ready to stream as a download
    """
    wb = openpyxl.Workbook()
    input_data = result.get_input_data()
    efficiency = result.predicted_efficiency
    severity = _get_severity(efficiency)
    hex_col = SEVERITY_HEX[severity]

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'

    # Helpers for consistent styling
    def _header_fill(hex_color: str) -> PatternFill:
        return PatternFill('solid', fgColor=hex_color)

    def _thin_border() -> Border:
        s = Side(style='thin', color='1E293B')
        return Border(left=s, right=s, top=s, bottom=s)

    # Title row
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = 'Solar Panel Degradation Predictor — Prediction Report'
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = _header_fill(THEME_COLORS_HEX['navy'])
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Metadata rows
    meta = [
        ('Result ID', f'#{result.id}'),
        ('Generated', result.created_at.strftime('%Y-%m-%d %H:%M:%S UTC')),
        ('Predicted Efficiency', f'{efficiency:.4f}  ({efficiency * 100:.1f}%)'),
        ('Severity', severity.upper()),
        ('Maintenance Recommendation', result.maintenance_recommendation),
    ]
    for row_idx, (label, value) in enumerate(meta, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=10, color=THEME_COLORS_HEX['amber'])
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.font = Font(size=10)
        # Colour-code the efficiency and severity rows
        if label in ('Predicted Efficiency', 'Severity', 'Maintenance Recommendation'):
            val_cell.font = Font(bold=True, size=10, color=hex_col)
        if label in ('Predicted Efficiency', 'Severity'):
            val_cell.fill = _header_fill(SEVERITY_TINT_HEX[severity])
        for col in range(1, 3):
            ws.cell(row=row_idx, column=col).border = _thin_border()

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 55

    # Model note (below metadata)
    note_row = len(meta) + 3
    ws.merge_cells(f'A{note_row}:D{note_row}')
    note_cell = ws[f'A{note_row}']
    note_cell.value = (
        'Model: Gradient Boosting (R2=0.790, RMSE=0.048). '
        'Top feature: irradiance (importance 0.674). '
        'Thresholds from EDA (dataset mean ~0.52).'
    )
    note_cell.font = Font(italic=True, size=9, color=THEME_COLORS_HEX['muted'])
    note_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    # ── Sheet 2: Input Data ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Input Data')

    # Header row
    features = list(input_data.keys())
    values = list(input_data.values())

    for col_idx, feat in enumerate(features, start=1):
        hdr = ws2.cell(row=1, column=col_idx, value=feat)
        hdr.font = Font(bold=True, color='FFFFFF')
        hdr.fill = _header_fill(THEME_COLORS_HEX['navy'])
        hdr.alignment = Alignment(horizontal='center')
        hdr.border = _thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(12, len(feat) + 2)

    # Value row
    for col_idx, val in enumerate(values, start=1):
        val_str = f'{val:.4f}' if isinstance(val, float) else str(val)
        cell = ws2.cell(row=2, column=col_idx, value=val_str)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = _header_fill(THEME_COLORS_HEX['tint_row_a'])
        cell.border = _thin_border()

    # Keep alternating light row shading for readability if more rows are added later.
    for row_idx in range(3, ws2.max_row + 1):
        fill = THEME_COLORS_HEX['tint_row_a'] if row_idx % 2 == 0 else THEME_COLORS_HEX['tint_row_b']
        for col_idx in range(1, ws2.max_column + 1):
            ws2.cell(row=row_idx, column=col_idx).fill = _header_fill(fill)

    # Freeze header
    ws2.freeze_panes = 'A2'

    # Output to bytes buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
